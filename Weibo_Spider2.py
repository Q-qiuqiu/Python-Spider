import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import jieba
from wordcloud import WordCloud
import matplotlib.pyplot as plt
from collections import Counter

# 停用词列表
stopwords = {"的", "了", "在", "是", "我", "和", "也", "有", "不", "人", "这", "中", "大", "为", "就", "都", "而", "上",
             "到", "说", "会", "与", "去", "可以", "将", "没有", "它", "你", "要", "我们", "他", "她", "呢", "对",
             "但是"}

# 爬取数据
def fetch_data(user_id, max_posts):
    base_url = f"https://weibo.cn/{user_id}"
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:133.0) Gecko/20100101 Firefox/133.0',
        'Cookie': 'SCF=AtDLxZdp81N2zZ-Oidxa9Voc0hK9JLOUeivKScT5CyzK_A-g-jdq9GGO_m92KBN3riQ9d-Io14xnv5YU6cDZlRk.; SUB=_2A25KYEXxDeRhGeFG7VoY8izJwziIHXVpHMc5rDV6PUJbktANLRnBkW1NeR2a9DQxzNpXUBnDJrJvPrTQJz2t-9Lv; SUBP=0033WrSXqPxfM725Ws9jqgMF55529P9D9W59cYickyxv0avrWQyUnBGh5NHD95QN1hqR1KzESKnXWs4DqcjMi--NiK.Xi-2Ri--ciKnRi-zNS0nc1h.Eeo-RS5tt; ALF=1737212577; _T_WM=d176226a71ae60b89d7218d78df0d425',
    }
    posts = []
    page = 1
    total_posts = 0

    while total_posts < max_posts:
        response = requests.get(f"{base_url}?page={page}", headers=headers)
        print(f"Fetching page {page}, Status Code: {response.status_code}")

        if response.status_code != 200:
            print(f"Failed to fetch page {page}, status code: {response.status_code}")
            break

        soup = BeautifulSoup(response.text, 'lxml')
        post_divs = soup.find_all('div', class_='c', id=True)

        if not post_divs:
            print("No more posts available.")
            break

        for post in post_divs:
            content_tag = post.find('span', class_='ctt')
            time_tag = post.find('span', class_='ct')

            if content_tag and time_tag:
                content = content_tag.get_text(strip=True)
                time = time_tag.get_text(strip=True)

                comments = []
                posts.append({
                    'content': content,
                    'time': time,
                    'comments': comments
                })
                total_posts += 1

                if total_posts >= max_posts:
                    break
        page += 1
    return posts


def save_to_excel(data, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Weibo Posts"

    # 写入标题行
    ws.append(["Content", "Time", "Comments"])

    # 写入数据
    for post in data:
        comments_str = "\n".join(post['comments']) if post['comments'] else "No comments"
        ws.append([post['content'], post['time'], comments_str])

    # 自动调整列宽
    for col in range(1, 4):  # 3 列：Content, Time, Comments
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(filename)
    print(f"Data saved to {filename}")


def generate_wordcloud(data, user_name):
    text = ""
    for post in data:
        text += post['content'] + " "
        for comment in post['comments']:
            text += comment + " "

    words = jieba.cut(text)
    words = [word for word in words if word not in stopwords and len(word.strip()) > 1]

    if not words:
        print(f"No valid words for word cloud for {user_name}!")
        return

    word_counts = Counter(words)
    print(f"Top 15 most frequent words for {user_name}:", word_counts.most_common(15))

    font_path = "C:\Windows\Fonts\msyh.ttc"
    wordcloud = WordCloud(
        width=800,
        height=400,
        background_color='white',
        font_path=font_path
    ).generate_from_frequencies(word_counts)

    plt.figure(figsize=(10, 5))
    plt.imshow(wordcloud, interpolation='bilinear')
    plt.axis("off")
    plt.title(f"Word Cloud for {user_name}")
    plt.show()


if __name__ == "__main__":
    user_ids = {
        "Mixue": "1704709632",
        "Yanjin": "2909048750",
        "Songshu": "2559971200",
        "Taoli": "1658876261",
        "Liangpin": "1851770484"
    }

    max_posts = 100  # 每个用户爬取的微博条数

    for user_name, user_id in user_ids.items():
        print(f"Fetching Weibo data for {user_name}...")
        weibo_data = fetch_data(user_id, max_posts)

        if weibo_data:
            excel_filename = f"{user_name}_data.xlsx"
            print(f"Saving data for {user_name} to {excel_filename}...")
            save_to_excel(weibo_data, excel_filename)

            print(f"Generating word cloud for {user_name}...")
            generate_wordcloud(weibo_data, user_name)
        else:
            print(f"No data fetched for {user_name}.")
