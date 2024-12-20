import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import jieba
from wordcloud import WordCloud
import matplotlib.pyplot as plt
from collections import Counter

# 停用词列表
stopwords = {"的", "了", "在", "是", "我", "和", "也", "有", "不", "人", "这", "中", "大", "为", "就", "都", "而", "上",
             "到", "说", "会", "与", "去", "可以", "将", "没有", "它", "你", "要", "我们", "他", "她", "呢", "对",
             "但是"}

#爬取数据
def fetch_data(user_id, max_posts):
    base_url = f"https://weibo.cn/{user_id}"
    # Cookie可能会过期
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:133.0) Gecko/20100101 Firefox/133.0',
        'Cookie': 'SCF=AtDLxZdp81N2zZ-Oidxa9Voc0hK9JLOUeivKScT5CyzK_A-g-jdq9GGO_m92KBN3riQ9d-Io14xnv5YU6cDZlRk.; SUB=_2A25KYEXxDeRhGeFG7VoY8izJwziIHXVpHMc5rDV6PUJbktANLRnBkW1NeR2a9DQxzNpXUBnDJrJvPrTQJz2t-9Lv; SUBP=0033WrSXqPxfM725Ws9jqgMF55529P9D9W59cYickyxv0avrWQyUnBGh5NHD95QN1hqR1KzESKnXWs4DqcjMi--NiK.Xi-2Ri--ciKnRi-zNS0nc1h.Eeo-RS5tt; ALF=1737212577; _T_WM=d176226a71ae60b89d7218d78df0d425',
    }
    posts = []
    page = 1
    total_posts = 0

    while total_posts < max_posts:
        response = requests.get(f"{base_url}?page={page}", headers=headers)
        print(f"Fetching page {page}, Status Code: {response.status_code}")

        if response.status_code != 200:
            print(f"Failed to fetch page {page}, status code: {response.status_code}")
            break

        soup = BeautifulSoup(response.text, 'lxml')
        post_divs = soup.find_all('div', class_='c', id=True)

        if not post_divs:
            print("No more posts available.")
            break

        for post in post_divs:
            content_tag = post.find('span', class_='ctt')
            time_tag = post.find('span', class_='ct')

            if content_tag and time_tag:
                content = content_tag.get_text(strip=True)
                time = time_tag.get_text(strip=True)

                # 提取评论链接
                comment_link_tag = post.find('a', class_='cc')
                comment_url = comment_link_tag['href'] if comment_link_tag else None
                print(f"Comment URL: {comment_url}")  # 打印评论 URL

                comments = []
                if comment_url:
                    # 从评论 URL 中提取帖子 ID
                    post_id = comment_url.split('/')[-1].split('?')[0]
                    comment_page_url = f"https://weibo.cn/comment/{post_id}?uid={user_id}&rl=0"

                    comment_response = requests.get(comment_page_url, headers=headers)
                    comment_soup = BeautifulSoup(comment_response.text, 'lxml')
                    comment_divs = comment_soup.find_all('div', class_='c')

                    # 获取最多 10 条评论，并从第二条开始
                    for comment in comment_divs[5:15]:  # 跳过正文和部分评论，最多获取 10 条
                        comment_content_tag = comment.find('span', class_='ctt')
                        if comment_content_tag:
                            comments.append(comment_content_tag.get_text(strip=True))

                posts.append({
                    'content': content,
                    'time': time,
                    'comments': comments
                })
                total_posts += 1

                if total_posts >= max_posts:
                    break
        page += 1
    return posts


def save_to_excel(data, filename="weibo_posts.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Weibo Posts"

    # 写入标题行
    ws.append(["Content", "Time", "Comments"])

    # 写入数据
    for post in data:
        comments_str = "\n".join(post['comments']) if post['comments'] else "No comments"
        ws.append([post['content'], post['time'], comments_str])

    # 自动调整列宽
    for col in range(1, 4):  # 3 列：Content, Time, Comments
        max_length = 0
        column = get_column_letter(col)
        for cell in ws[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)  # 加 2 用于填充
        ws.column_dimensions[column].width = adjusted_width

    # 自动调整行高（根据内容调整行高，尤其是评论列）
    for row in ws.iter_rows():
        max_height = 0
        for cell in row:
            try:
                if isinstance(cell.value, str):  # 如果是字符串内容
                    max_height = max(max_height, len(cell.value.split("\n")))
            except:
                pass
        if max_height > 1:
            ws.row_dimensions[cell.row].height = max_height * 15  # 每行高 15px 乘以评论行数

    wb.save(filename)
    print(f"Data saved to {filename}")


def generate_wordcloud(data):
    # 提取所有文本内容（帖子内容 + 评论）
    text = ""
    for post in data:
        text += post['content'] + " "
        for comment in post['comments']:
            text += comment + " "

    # 使用分词处理文本
    words = jieba.cut(text)
    words = [word for word in words if word not in stopwords and len(word.strip()) > 1]  # 去除停用词并且只保留长度大于1的词

    # 如果分词后的文本为空，返回提示
    if not words:
        print("No valid words for word cloud!")
        return

    # 统计词频
    word_counts = Counter(words)
    print("Top 10 most frequent words:", word_counts.most_common(15))  # 打印前15个频率最高的词

    # 生成词云图
    font_path = "C:\Windows\Fonts\msyh.ttc"  # 替换为中文字体文件路径
    wordcloud = WordCloud(
        width=800,
        height=400,
        background_color='white',
        font_path=font_path  # 指定中文字体
    ).generate_from_frequencies(word_counts)

    # 绘制词云图
    plt.figure(figsize=(10, 5))
    plt.imshow(wordcloud, interpolation='bilinear')
    plt.axis("off")
    plt.show()

if __name__ == "__main__":
    mixue_id = "1704709632"  # 蜜雪冰城 ID
    yanjin_id=  "2909048750" # 盐津铺子 ID
    songshu_id = "2559971200"  # 三只松鼠 ID
    taoli_id = "1658876261"  # 桃李面包 ID
    user_id = "1851770484"  # 良品铺子 ID

    max_posts = 100  # 要爬取的微博条数
    print("Fetching Weibo data...")
    weibo_data = fetch_data(user_id, max_posts)

    if weibo_data:
        print("Saving data to Excel...")
        save_to_excel(weibo_data)

        print("Generating word cloud...")
        generate_wordcloud(weibo_data)
    else:
        print("No data fetched.")
